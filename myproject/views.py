from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q, Count, Avg, Prefetch, Sum
from datetime import datetime, date
import csv
import json
from .forms import (
    RegisterForm, ClassForm, StudentForm, AttendanceForm, ScoreForm,
    GradeCalculationSettingsForm, CSVUploadForm, ProfileUpdateForm
)
from .models import (
    User, Class, Student, GradeSummary, Setting, UserSettings,
    Attendance, Score, GradeCalculationSettings, ActivityLog,
    GradeCategory, GradeItem, StudentScore, TransmutationTable,
    EmailVerification
)
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import csv
import io
import openpyxl
import unicodedata 
from django.db import transaction

# -----------------------------
# HELPER FUNCTIONS
# -----------------------------
def get_current_school_year(user):
    """Get the current school year setting for the user"""
    try:
        setting, created = Setting.objects.get_or_create(
            user=user,
            defaults={'school_year': '25-1', 'theme': 'light'}
        )
        return setting.school_year or '25-1'
    except Exception:
        return "25-1"

def get_user_theme(user):
    """Get the current theme setting for the user"""
    try:
        return user.app_setting.theme
    except:
        return "light"  # Default fallback

# -----------------------------
# EMAIL VERIFICATION HELPER
# -----------------------------
def send_verification_email(email, code):
    """Send verification code to email"""
    try:
        subject = 'ASCReM Email Verification'
        message = f'Your verification code is: {code}\n\nThis code will expire in 10 minutes.'
        send_mail(
            subject,
            message,
            'noreply@ascrem.com',
            [email],
            fail_silently=False,
        )
        print(f"\n=== EMAIL SENT ===\nTo: {email}\nCode: {code}\n==================\n")
        return True
    except Exception as e:
        print(f"Email sending failed: {e}")
        return False


# -----------------------------
# AUTH / INDEX / DASHBOARD
# -----------------------------
def register(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Account created successfully! Please log in.")
            return redirect("index")
    else:
        form = RegisterForm()
    return render(request, "register.html", {"form": form})


@login_required
def dashboard_view(request):
    return render(request, "dashboard.html")


def index(request):
    register_form = RegisterForm()
    login_form = AuthenticationForm()
    active_tab = "login"  # default tab

    if request.method == "POST":

        # -----------------------------
        # Registration
        # -----------------------------
        if "register_submit" in request.POST:
            active_tab = "register"  # keep register tab active

            register_form = RegisterForm(request.POST)
            if register_form.is_valid():
                password1 = register_form.cleaned_data.get("password1")
                password2 = register_form.cleaned_data.get("password2")
                email = register_form.cleaned_data.get("email")

                if len(password1) != 4 or len(password2) != 4:
                    messages.error(request, "Password must be exactly 4 characters long.")
                elif password1 != password2:
                    messages.error(request, "Passwords do not match.")
                else:
                    # Generate and send verification code
                    code = EmailVerification.generate_code()
                    EmailVerification.objects.filter(email=email).delete()
                    EmailVerification.objects.create(email=email, code=code)
                    
                    if send_verification_email(email, code):
                        request.session['pending_registration'] = {
                            'instructor_id': register_form.cleaned_data.get('instructor_id'),
                            'username': register_form.cleaned_data.get('username'),
                            'email': email,
                            'first_name': register_form.cleaned_data.get('first_name'),
                            'last_name': register_form.cleaned_data.get('last_name'),
                            'password': password1,
                        }
                        messages.success(request, f"Verification code sent to {email}. Please check your email and enter the code below.")
                        return render(request, "index.html", {
                            "register_form": register_form,
                            "login_form": login_form,
                            "active_tab": active_tab,
                            "show_verification_modal": True
                        })
                    else:
                        messages.error(request, "Failed to send verification email. Please try again.")
            else:
                # Show single consolidated error message
                messages.error(request, "Registration failed. Please check your information and try again.")
                return render(request, "index.html", {
                    "register_form": register_form,
                    "login_form": login_form,
                    "active_tab": active_tab
                })

        # -----------------------------
        # Login
        # -----------------------------
        elif "login_submit" in request.POST:
            active_tab = "login"

            username = request.POST.get("username")
            password = request.POST.get("password")

            if len(password) != 4:
                messages.error(request, "Password must be exactly 4 characters long.")
            else:
                user = authenticate(request, username=username, password=password)
                if user:
                    login(request, user)
                    return redirect("dashboard")
                else:
                    messages.error(request, "Invalid Instructor_ID or password.")

    return render(request, "index.html", {
        "register_form": register_form,
        "login_form": login_form,
        "active_tab": active_tab
    })


def verify_email(request):
    """Handle email verification"""
    if request.method == "POST":
        code = request.POST.get('verification_code')
        pending_data = request.session.get('pending_registration')
        
        if not pending_data:
            messages.error(request, 'No pending registration found.')
            return JsonResponse({'success': False})
        
        try:
            verification = EmailVerification.objects.get(
                email=pending_data['email'],
                code=code,
                is_verified=False
            )
            
            if verification.is_expired():
                messages.error(request, 'Verification code has expired.')
                return JsonResponse({'success': False})
            
            # Create user account
            user = User.objects.create_user(
                username=pending_data['username'],
                instructor_id=pending_data['instructor_id'],
                email=pending_data['email'],
                first_name=pending_data['first_name'],
                last_name=pending_data['last_name'],
                password=pending_data['password']
            )
            
            # Create related objects
            Setting.objects.create(user=user)
            UserSettings.objects.create(user=user)
            
            # Mark verification as complete
            verification.is_verified = True
            verification.save()
            
            # Clear session data
            del request.session['pending_registration']
            
            messages.success(request, 'Account created successfully! Please log in.')
            return JsonResponse({'success': True})
            
        except EmailVerification.DoesNotExist:
            messages.error(request, 'Invalid verification code.')
            return JsonResponse({'success': False})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
def dashboard(request):
    # Get current school year
    current_school_year = get_current_school_year(request.user)
    
    # Basic statistics - filtered by school year
    total_classes = Class.objects.filter(instructor=request.user, school_year=current_school_year).count()
    total_students = Student.objects.filter(class_obj__instructor=request.user, class_obj__school_year=current_school_year).distinct().count()
    pending_grades = GradeSummary.objects.filter(class_obj__instructor=request.user, class_obj__school_year=current_school_year, is_locked=False).count()

    # Recent activities
    recent_activities = ActivityLog.objects.filter(user=request.user)[:10]

    # Top performing students - filtered by school year
    top_students = GradeSummary.objects.filter(
        class_obj__instructor=request.user,
        class_obj__school_year=current_school_year
    ).order_by('-final_grade')[:5]

    # Attendance statistics - filtered by school year
    total_attendance_records = Attendance.objects.filter(class_obj__instructor=request.user, class_obj__school_year=current_school_year).count()
    present_count = Attendance.objects.filter(class_obj__instructor=request.user, class_obj__school_year=current_school_year, status='Present').count()
    absent_count = Attendance.objects.filter(class_obj__instructor=request.user, class_obj__school_year=current_school_year, status='Absent').count()
    late_count = Attendance.objects.filter(class_obj__instructor=request.user, class_obj__school_year=current_school_year, status='Late').count()
    excused_count = Attendance.objects.filter(class_obj__instructor=request.user, class_obj__school_year=current_school_year, status='Excused').count()

    # Calculate attendance percentage
    attendance_percentage = (present_count / total_attendance_records * 100) if total_attendance_records > 0 else 0

    # Class distribution by semester - filtered by school year
    class_distribution = Class.objects.filter(instructor=request.user, school_year=current_school_year).values('semester').annotate(count=Count('id'))

    # Recent classes - filtered by school year
    recent_classes = Class.objects.filter(instructor=request.user, school_year=current_school_year).order_by('-id')[:5]

    # Students with low attendance (for dropping list) - filtered by school year
    dropping_list = []
    for class_obj in Class.objects.filter(instructor=request.user, school_year=current_school_year):
        students = Student.objects.filter(class_obj=class_obj)
        for student in students:
            total_days = Attendance.objects.filter(student=student, class_obj=class_obj).count()
            absent_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Absent').count()
            if total_days > 0:
                absence_rate = (absent_days / total_days) * 100
                if absence_rate > 20:  # More than 20% absence
                    dropping_list.append({
                        'student': student,
                        'class': class_obj,
                        'absence_rate': round(absence_rate, 2)
                    })

    # Sort by absence rate
    dropping_list.sort(key=lambda x: x['absence_rate'], reverse=True)
    dropping_list = dropping_list[:10]  # Top 10

    context = {
        "total_classes": total_classes,
        "total_students": total_students,
        "pending_grades": pending_grades,
        "recent_activities": recent_activities,
        "top_students": top_students,
        "total_attendance_records": total_attendance_records,
        "present_count": present_count,
        "absent_count": absent_count,
        "late_count": late_count,
        "excused_count": excused_count,
        "attendance_percentage": round(attendance_percentage, 2),
        "class_distribution": list(class_distribution),
        "recent_classes": recent_classes,
        "dropping_list": dropping_list,
        "current_school_year": current_school_year,
    }
    
    context['user_theme'] = get_user_theme(request.user)
    return render(request, "dashboard.html", context)

def logout_view(request):
    logout(request)
    return redirect("index")


def instructor_panel(request):
    return render(request, "instructor.html")




@login_required
def class_detail(request, class_id):
    selected_class = get_object_or_404(Class, id=class_id, instructor=request.user)
    students = Student.objects.filter(class_obj=selected_class)

    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save(commit=False)
            student.class_obj = selected_class
            student.save()
            messages.success(request, "Student added successfully!")
            return redirect("class_detail", class_id=class_id)
    else:
        form = StudentForm()

    current_school_year = get_current_school_year(request.user)
    return render(request, "class.html", {
        "selected_class": selected_class,
        "students": students,
        "student_form": form,
        "class_form": ClassForm(user=request.user),
        "classes": Class.objects.filter(instructor=request.user, school_year=current_school_year),
    })


@login_required
def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id, class_obj__instructor=request.user)
    return render(request, "student_detail.html", {"student": student})


@login_required
def add_class(request):
    if request.method == "POST":
        form = ClassForm(request.POST, user=request.user)
        if form.is_valid():
            new_class = form.save(commit=False)
            new_class.instructor = request.user
            new_class.save()
            messages.success(request, "Class added successfully!")
    return redirect("class_panel")


@login_required
def edit_class(request, class_id):
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    if request.method == "POST":
        form = ClassForm(request.POST, instance=class_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Class updated successfully!")
            return redirect("class_panel")
    else:
        form = ClassForm(instance=class_obj)
    return render(request, "edit_class.html", {"form": form})


@login_required
def delete_class(request, class_id):
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    class_obj.delete()
    messages.success(request, "Class deleted successfully!")
    return redirect("class_panel")


@login_required
def add_student(request, class_id):
    selected_class = get_object_or_404(Class, id=class_id, instructor=request.user)
    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save(commit=False)
            student.class_obj = selected_class
            student.save()
            messages.success(request, "Student added successfully!")
    return redirect("class_detail", class_id=class_id)


@login_required
def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id, class_obj__instructor=request.user)
    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, "Student updated successfully!")
            return redirect("class_detail", class_id=student.class_obj.id)
    else:
        form = StudentForm(instance=student)
    return render(request, "edit_student.html", {"form": form})


@login_required
def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id, class_obj__instructor=request.user)
    class_id = student.class_obj.id
    student.delete()
    messages.success(request, "Student deleted successfully!")
    return redirect("class_detail", class_id=class_id)


def edit_class(request, class_id):
    cls = get_object_or_404(Class, id=class_id)

    if request.method == "POST":
        form = ClassForm(request.POST, instance=cls)
        if form.is_valid():
            form.save()
            return redirect("class_list")  # or wherever you want to go after saving
    else:
        form = ClassForm(instance=cls)

    return render(request, "edit_class.html", {"form": form, "class_id": class_id})


def class_list(request):
    classes = Class.objects.all()
    return render(request, "class_list.html", {"classes": classes})


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Prefetch
from .models import Class, Student, GradeCategory, GradeItem, StudentScore, GradeSummary, GradeCalculationSettings, TransmutationTable, ActivityLog


@login_required
def grades_panel(request):
    """
    Unified grades_panel:
    - GET: show classes, selected class, selected category, categories + items + student scores, summaries, settings
    - POST: perform actions: add/edit/delete category/item, save_scores, compute_final
    """

    # -------------------------
    # Filter Options
    # -------------------------
    current_school_year = get_current_school_year(request.user)
    selected_school_year = current_school_year  # Always use current school year
    selected_semester = request.GET.get("semester")
    selected_class_id = request.GET.get("class_id") or request.POST.get("selected_class_id")
    selected_category_id = request.GET.get("category_id") or request.POST.get("selected_category_id")

    semesters = (
        Class.objects.filter(instructor=request.user, school_year=current_school_year)
        .values_list('semester', flat=True)
        .distinct()
        .order_by('semester')
    )

    # -------------------------
    # Base Query for Classes - default to current school year
    # -------------------------
    classes = Class.objects.filter(instructor=request.user, school_year=selected_school_year)
    if selected_semester:
        classes = classes.filter(semester=selected_semester)
    classes = classes.order_by('program')

    # -------------------------
    # Initialize Variables
    # -------------------------
    selected_class = None
    selected_category = None
    students = []
    summaries = {}
    grade_settings = None
    categories = []
    transmutation = []
    score_map = {}
    category_totals = {}
    students_qs = []

    # -------------------------
    # Helper Functions
    # -------------------------
    def load_transmutation():
        rows = list(TransmutationTable.objects.all().order_by('-min_percentage'))
        if rows:
            return rows
        # Default fallback
        class T: pass
        default = [
            (96, 100, 1.0),
            (90, 95, 1.25),
            (85, 89, 1.5),
            (80, 84, 1.75),
            (75, 79, 2.0),
            (0, 74.99, 5.0),
        ]
        out = []
        for mn, mx, eq in default:
            t = T()
            t.min_percentage = mn
            t.max_percentage = mx
            t.equivalent_grade = eq
            out.append(t)
        return out

    def get_equivalent_grade(score, total_items=None):
        # If total_items provided, compute scale (for individual items)
        tables = {
            10: [(i, round(4.0-(i*3.0/10),1)) for i in range(11)],
            20: [(i, round(4.0-(i*3.0/20),1)) for i in range(21)],
            30: [(i, round(4.0-(i*3.0/30),1)) for i in range(31)],
            40: [(i, round(4.0-(i*3.0/40),1)) for i in range(41)],
            50: [(i, round(4.0-(i*3.0/50),1)) for i in range(51)],
            100: [(i, round(4.0-(i*3.0/100),1)) for i in range(101)],
        }
        if total_items and total_items in tables:
            table = tables[total_items]
            closest = min(table, key=lambda x: abs(x[0]-score))
            return closest[1]
        else:
            percentage = score
            rows = load_transmutation()
            for r in rows:
                if r.min_percentage <= percentage <= r.max_percentage:
                    return r.equivalent_grade
            return rows[-1].equivalent_grade

    def get_final_equivalent_grade(percentage):
        """Convert final grade percentage to equivalent grade (1.0-5.0 scale)"""
        # Based on the grading table provided
        grade_scale = [
            (90, 100, 1.0), (89, 89, 1.1), (88, 88, 1.2), (87, 87, 1.3), (86, 86, 1.4),
            (85, 85, 1.5), (84, 84, 1.6), (83, 83, 1.7), (82, 82, 1.8), (81, 81, 1.9),
            (80, 80, 2.0), (79, 79, 2.1), (78, 78, 2.2), (77, 77, 2.3), (76, 76, 2.4),
            (75, 75, 2.5), (74, 74, 2.6), (73, 73, 2.7), (72, 72, 2.8), (71, 71, 2.9),
            (70, 70, 3.0), (69, 69, 3.1), (68, 68, 3.2), (67, 67, 3.3), (66, 66, 3.4),
            (65, 65, 3.5), (64, 64, 3.6), (63, 63, 3.7), (62, 62, 3.8), (61, 61, 3.9),
            (60, 60, 4.0), (59, 59, 4.1), (58, 58, 4.2), (57, 57, 4.3), (56, 56, 4.4),
            (55, 55, 4.5), (54, 54, 4.6), (53, 53, 4.7), (52, 52, 4.8), (0, 51, 4.9),
        ]
        
        percentage = round(percentage)  # Round to nearest integer
        
        for min_pct, max_pct, equiv in grade_scale:
            if min_pct <= percentage <= max_pct:
                return equiv
        
        return 5.0  # Default for very low scores

    # -------------------------
    # Handle POST Actions
    # -------------------------
    if request.method == "POST":
        action = request.POST.get("action")
        if selected_class_id:
            selected_class = get_object_or_404(Class, id=selected_class_id, instructor=request.user)
            students_qs = Student.objects.filter(class_obj=selected_class)

        # ---------- CATEGORY CRUD ----------
        if action == "add_category":
            name = request.POST.get("category_name", "").strip()
            percentage = float(request.POST.get("category_percentage", 0) or 0)
            if not name:
                messages.error(request, "Category name cannot be empty.")
                return redirect(f"{request.path}?class_id={selected_class.id}")

            existing_total = GradeCategory.objects.filter(class_obj=selected_class).aggregate(
                total=Sum('percentage')
            )['total'] or 0

            if existing_total + percentage > 100:
                messages.error(request, f"Cannot add category '{name}'. Total percentage would exceed 100% (Current total: {existing_total}%).")
                return redirect(f"{request.path}?class_id={selected_class.id}")

            GradeCategory.objects.create(class_obj=selected_class, name=name, percentage=percentage)
            messages.success(request, f"Category '{name}' added successfully!")
            return redirect(f"{request.path}?class_id={selected_class.id}")

        elif action == "edit_category":
            cat_id = request.POST.get("category_id")
            cat = get_object_or_404(GradeCategory, id=cat_id, class_obj__instructor=request.user)
            new_name = request.POST.get("category_name", "").strip()
            new_percentage = float(request.POST.get("category_percentage", 0) or 0)

            other_total = GradeCategory.objects.filter(class_obj=selected_class).exclude(id=cat.id).aggregate(
                total=Sum('percentage')
            )['total'] or 0

            if other_total + new_percentage > 100:
                messages.error(request, f"Cannot update category '{new_name}'. Total percentage would exceed 100% (Current total excluding this category: {other_total}%).")
                return redirect(f"{request.path}?class_id={selected_class.id}")

            cat.name = new_name
            cat.percentage = new_percentage
            cat.save()
            messages.success(request, f"Category '{new_name}' updated successfully!")
            return redirect(f"{request.path}?class_id={selected_class.id}")

        elif action == "delete_category":
            cat_id = request.POST.get("category_id")
            cat = get_object_or_404(GradeCategory, id=cat_id, class_obj__instructor=request.user)
            cat.delete()
            messages.success(request, "Category deleted successfully!")
            return redirect(f"{request.path}?class_id={selected_class.id}")

        # ---------- ITEM CRUD ----------
        elif action == "add_item":
            category_id = request.POST.get("category_id")
            item_name = request.POST.get("item_name", "").strip()
            total_items = int(request.POST.get("total_items", 0) or 0)
            passing_percentage = float(request.POST.get("passing_percentage", 0) or 0)
            category = get_object_or_404(GradeCategory, id=category_id, class_obj__instructor=request.user)
            if item_name:
                GradeItem.objects.create(category=category, item_name=item_name, total_items=total_items, passing_percentage=passing_percentage)
            return redirect(f"{request.path}?class_id={selected_class.id}&category_id={category.id}")

        elif action == "edit_item":
            item_id = request.POST.get("item_id")
            item = get_object_or_404(GradeItem, id=item_id, category__class_obj__instructor=request.user)
            item.item_name = request.POST.get("item_name", "").strip()
            item.total_items = int(request.POST.get("total_items", 0) or 0)
            item.passing_percentage = float(request.POST.get("passing_percentage", 0) or 0)
            item.save()
            return redirect(f"{request.path}?class_id={selected_class.id}&category_id={item.category.id}")

        elif action == "delete_item":
            item_id = request.POST.get("item_id")
            item = get_object_or_404(GradeItem, id=item_id, category__class_obj__instructor=request.user)
            category_id = item.category.id
            item.delete()
            return redirect(f"{request.path}?class_id={selected_class.id}&category_id={category_id}")

        # ---------- SAVE SCORES ----------
        elif action == "save_scores":
            saved_count = 0
            with transaction.atomic():
                for key, value in request.POST.items():
                    if key.startswith("scores-"):
                        try:
                            _, student_id, item_id = key.split("-")
                            if value.strip() == "":
                                continue
                            score_value = float(value)
                            student = Student.objects.get(id=student_id, class_obj=selected_class)
                            item = GradeItem.objects.get(id=item_id, category__class_obj=selected_class)
                            StudentScore.objects.update_or_create(
                                student=student,
                                item=item,
                                defaults={"score_percentage": score_value}
                            )
                            saved_count += 1
                        except Exception as e:
                            print("Error saving score:", e)
                            continue
            messages.success(request, f"Saved {saved_count} scores successfully!")
            return redirect(f"{request.path}?class_id={selected_class.id}&category_id={selected_category_id or ''}")

        # ---------- COMPUTE FINAL GRADES ----------
        elif action == "compute_final":
            categories_qs = GradeCategory.objects.filter(class_obj=selected_class)
            total_percentage = sum([c.percentage for c in categories_qs])
            if abs(total_percentage - 100.0) > 0.001:
                messages.error(request, "Total category percentage must equal 100%.")
                return redirect(f"{request.path}?class_id={selected_class.id}")

            with transaction.atomic():
                try:
                    passing_threshold = selected_class.grade_settings.passing_grade
                except GradeCalculationSettings.DoesNotExist:
                    passing_threshold = 75.0

                for student in students_qs:
                    final_percentage = 0.0
                    for category in categories_qs:
                        items = GradeItem.objects.filter(category=category)
                        item_percentages = []
                        for item in items:
                            ss = StudentScore.objects.filter(student=student, item=item).first()
                            if ss and item.total_items > 0:
                                # Convert raw score to percentage
                                item_pct = (ss.score_percentage / item.total_items) * 100
                            else:
                                item_pct = 0.0
                            item_percentages.append(item_pct)

                        # Calculate category average from percentages
                        category_average = sum(item_percentages) / len(item_percentages) if item_percentages else 0.0
                        
                        # Apply category weight
                        weighted_category = category_average * (category.percentage / 100)
                        final_percentage += weighted_category

                    final_percentage = round(final_percentage, 2)
                    remarks = "Passed" if final_percentage >= passing_threshold else "Failed"
                    
                    # Calculate equivalent grade
                    equivalent_grade = get_final_equivalent_grade(final_percentage)

                    GradeSummary.objects.update_or_create(
                        student=student,
                        class_obj=selected_class,
                        defaults={
                            "final_grade": final_percentage,
                            "equivalent_grade": equivalent_grade,
                            "remarks": remarks,
                            "is_locked": False,
                        }
                    )

            messages.success(request, "Final grades computed successfully!")
            return redirect(f"{request.path}?class_id={selected_class.id}")

    # -------------------------
    # GET: Prepare Data
    # -------------------------
    if selected_class_id:
        selected_class = Class.objects.filter(id=selected_class_id, instructor=request.user).first()
        if selected_class:
            students = Student.objects.filter(class_obj=selected_class).order_by('last_name', 'first_name')

            # Summaries - attach to each student object
            summaries_qs = GradeSummary.objects.filter(class_obj=selected_class).select_related('student').order_by('-final_grade')
            summaries = {s.student.id: s for s in summaries_qs}
            
            # Attach summary to each student for easy template access
            for student in students:
                student.summary = summaries.get(student.id)

            # Grade Settings
            try:
                grade_settings = selected_class.grade_settings
            except GradeCalculationSettings.DoesNotExist:
                grade_settings = None

            # Categories and Items
            categories = GradeCategory.objects.filter(class_obj=selected_class).order_by('name').prefetch_related(
                Prefetch('items', queryset=GradeItem.objects.all().order_by('id'))
            )

            # Score Map
            score_map = {}
            scores_qs = StudentScore.objects.filter(item__category__class_obj=selected_class).select_related('item', 'student')
            for s in scores_qs:
                score_map.setdefault(s.student.id, {})[s.item.id] = s

            # Category totals
            category_totals = {}
            for category in categories:
                category_totals[category.id] = {}
                items = category.items.all()
                for student in students:
                    total_score = 0.0
                    count_items = 0
                    for item in items:
                        ss = score_map.get(student.id, {}).get(item.id)
                        if ss:
                            # Convert raw score to percentage for display
                            item_percentage = (ss.score_percentage / item.total_items) * 100 if item.total_items > 0 else 0
                            total_score += item_percentage
                            count_items += 1
                            equivalent = get_equivalent_grade(ss.score_percentage, item.total_items)
                            setattr(ss, "equivalent_grade", equivalent)
                    category_average = (total_score / count_items) if count_items else 0.0
                    category_totals[category.id][student.id] = round(category_average, 2)

            # Selected Category
            if selected_category_id:
                selected_category = GradeCategory.objects.filter(id=selected_category_id, class_obj__instructor=request.user).first()

            transmutation = load_transmutation()
        else:
            messages.error(request, "Selected class not found or you don't have permission.")
            return redirect("grades_panel")

    # -------------------------
    # Stats
    # -------------------------
    passed_count = sum(1 for s in summaries.values() if s.remarks == 'Passed') if summaries else 0
    failed_count = sum(1 for s in summaries.values() if s.remarks == 'Failed') if summaries else 0
    excellent_count = sum(1 for s in summaries.values() if s.final_grade >= 90) if summaries else 0
    top_students = GradeSummary.objects.filter(class_obj__instructor=request.user).order_by('-final_grade')[:5]

    # Convert summaries to JSON-safe format
    import json
    summaries_json = json.dumps({
        str(student_id): {
            'final_grade': float(summary.final_grade),
            'remarks': summary.remarks
        }
        for student_id, summary in summaries.items()
    })

    # -------------------------
    # Render Template
    # -------------------------
    return render(request, "grades_summary.html", {
        "semesters": semesters,
        "selected_school_year": selected_school_year,
        "selected_semester": selected_semester,
        "classes": classes,
        "selected_class": selected_class,
        "students": students,
        "summaries": summaries,
        "summaries_json": summaries_json,
        "grade_settings": grade_settings,
        "categories": categories,
        "selected_category": selected_category,
        "score_map": score_map,
        "category_totals": category_totals,
        "transmutation": transmutation,
        "passed_count": passed_count,
        "failed_count": failed_count,
        "excellent_count": excellent_count,
        "top_students": top_students,
        "current_school_year": current_school_year,
    })

@login_required
def debug_grades(request):
    """Temporary debug view to check grade calculation"""
    
    # Get a specific class (replace with your class ID)
    class_id = request.GET.get('class_id')
    if not class_id:
        return HttpResponse("Please provide ?class_id=X in URL")
    
    selected_class = get_object_or_404(Class, id=class_id, instructor=request.user)
    
    debug_info = []
    debug_info.append(f"<h2>Debug Info for Class: {selected_class.program}</h2>")
    
    # Check categories
    categories = GradeCategory.objects.filter(class_obj=selected_class)
    debug_info.append(f"<h3>Categories ({categories.count()}):</h3>")
    debug_info.append("<ul>")
    for cat in categories:
        debug_info.append(f"<li>{cat.name} - Weight: {cat.percentage}%</li>")
    debug_info.append("</ul>")
    
    # Check total percentage
    total_percentage = sum([c.percentage for c in categories])
    debug_info.append(f"<p><strong>Total Category Weight: {total_percentage}%</strong></p>")
    
    # Check one student's scores
    students = Student.objects.filter(class_obj=selected_class)[:1]
    if students.exists():
        student = students.first()
        debug_info.append(f"<h3>Sample Student: {student.display_name}</h3>")
        
        for category in categories:
            debug_info.append(f"<h4>Category: {category.name} ({category.percentage}%)</h4>")
            items = GradeItem.objects.filter(category=category)
            debug_info.append(f"<p>Items in category: {items.count()}</p>")
            
            debug_info.append("<table border='1' cellpadding='5'>")
            debug_info.append("<tr><th>Item Name</th><th>Total Items</th><th>Student Score</th><th>Percentage</th></tr>")
            
            item_percentages = []
            for item in items:
                ss = StudentScore.objects.filter(student=student, item=item).first()
                if ss:
                    score = ss.score_percentage
                    percentage = (score / item.total_items) * 100 if item.total_items > 0 else 0
                    item_percentages.append(percentage)
                    debug_info.append(f"<tr><td>{item.item_name}</td><td>{item.total_items}</td><td>{score}</td><td>{percentage:.2f}%</td></tr>")
                else:
                    debug_info.append(f"<tr><td>{item.item_name}</td><td>{item.total_items}</td><td>NO SCORE</td><td>0.00%</td></tr>")
            
            debug_info.append("</table>")
            
            # Calculate category average
            if item_percentages:
                category_average = sum(item_percentages) / len(item_percentages)
                weighted_category = category_average * (category.percentage / 100)
                debug_info.append(f"<p><strong>Category Average: {category_average:.2f}%</strong></p>")
                debug_info.append(f"<p><strong>Weighted for Final Grade: {weighted_category:.2f}%</strong></p>")
            else:
                debug_info.append("<p><strong>No scores found for this category!</strong></p>")
    
        # Check if GradeSummary exists
        summary = GradeSummary.objects.filter(student=student, class_obj=selected_class).first()
        if summary:
            debug_info.append(f"<h3>GradeSummary Record:</h3>")
            debug_info.append(f"<p>Final Grade: {summary.final_grade}%</p>")
            debug_info.append(f"<p>Remarks: {summary.remarks}</p>")
        else:
            debug_info.append("<h3>NO GradeSummary record found!</h3>")
    
    return HttpResponse("<br>".join(debug_info))


# Add this to your urls.py temporarily:
# path('debug-grades/', views.debug_grades, name='debug_grades'),

@login_required
def attendance_panel(request):
    # Get all classes for the instructor - filtered by school year
    current_school_year = get_current_school_year(request.user)
    classes = Class.objects.filter(instructor=request.user, school_year=current_school_year)
    
    # Get selected class and date from GET parameters
    selected_class_id = request.GET.get('class_id')
    selected_date_str = request.GET.get('date')
    
    # Default to today if no date selected
    selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date() if selected_date_str else timezone.localdate()
    
    selected_class = None
    students = []
    attendance_records = []

    if selected_class_id:
        selected_class = get_object_or_404(Class, id=selected_class_id, instructor=request.user)
        
        # Fetch students using the related_name 'students'
        students = selected_class.students.all()
        


        if students.exists():
            # Ensure attendance exists for each student for the selected date
            for student in students:
                attendance, created = Attendance.objects.get_or_create(
                    class_obj=selected_class,
                    student=student,
                    date=selected_date,
                    defaults={'status': 'Present'}
                )

            # Fetch attendance records with related student data
            attendance_records = Attendance.objects.filter(
                class_obj=selected_class,
                date=selected_date
            ).select_related('student').order_by('student__last_name', 'student__first_name')
            


    return render(request, 'attendance.html', {
        'classes': classes,
        'selected_class': selected_class,
        'students': students,
        'attendance_records': attendance_records,
        'selected_date': selected_date,
    })

@login_required
def attendance_summary(request, class_id):
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)

    # Get all students who are enrolled OR have attendance records for this class
    students = Student.objects.filter(
        Q(enrollment__class_obj=class_obj) | 
        Q(attendance__class_obj=class_obj)
    ).distinct()

    attendance_stats = []
    for student in students:
        # Fetch all attendance records for this student in this class
        records = Attendance.objects.filter(student=student, class_obj=class_obj)

        total_days = records.count()
        present_days = records.filter(status='Present').count()
        absent_days = records.filter(status='Absent').count()
        late_days = records.filter(status='Late').count()
        excused_days = records.filter(status='Excused').count()

        attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0

        attendance_stats.append({
            'student': student,
            'total_days': total_days,
            'present_days': present_days,
            'absent_days': absent_days,
            'late_days': late_days,
            'excused_days': excused_days,
            'attendance_percentage': round(attendance_percentage, 2)
        })

    return render(request, 'attendance_summary.html', {
        'class_obj': class_obj,
        'attendance_stats': attendance_stats,
    })



@login_required
def update_attendance_ajax(request):
    """Handle AJAX POST for updating multiple attendance records"""
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            updates = data.get('updates', [])
            for item in updates:
                attendance_id = item.get('attendance_id')
                status = item.get('status')
                Attendance.objects.filter(id=attendance_id).update(status=status)
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Invalid request'})

@login_required
def update_attendance(request):
    """Update attendance status via AJAX safely"""
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            attendance_id = data.get('attendance_id')
            status = data.get('status')

            # Safely get the attendance record for the current instructor
            attendance = Attendance.objects.filter(
                id=attendance_id,
                class_obj__instructor=request.user
            ).first()

            if not attendance:
                return JsonResponse({'success': False, 'message': 'Attendance record not found or access denied'})

            # Update the status
            attendance.status = status
            attendance.save()

            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Updated attendance for {attendance.student.full_name}",
                description=f"Changed status to {status} on {attendance.date}",
                class_obj=attendance.class_obj,
                student=attendance.student
            )

            return JsonResponse({'success': True, 'message': 'Attendance updated successfully'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request'})




# -----------------------------
# PROFILE / SETTINGS / CSV / REPORTS (unchanged)
# -----------------------------
@login_required
def profile_view(request):
    """View and edit user profile"""
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('profile')
    else:
        form = ProfileUpdateForm(instance=request.user)

    return render(request, 'user_profile.html', {'form': form})


@login_required
def settings_view(request):
    """User settings page"""
    try:
        user_settings = request.user.app_setting
    except Setting.DoesNotExist:
        user_settings = Setting.objects.create(
            user=request.user,
            school_year='25-1',
            theme='light',
            notification_pref='email',
            default_report_format='pdf'
        )

    if request.method == 'POST':
        try:
            old_school_year = user_settings.school_year
            
            # Handle settings update
            user_settings.notification_pref = request.POST.get('notification_pref', 'email')
            user_settings.theme = request.POST.get('theme', 'light')
            user_settings.auto_generate_report = request.POST.get('auto_generate_report') == 'on'
            user_settings.default_report_format = request.POST.get('default_report_format', 'pdf')
            user_settings.school_year = request.POST.get('school_year', '25-1')
            user_settings.save()
            
            # Handle AJAX theme change
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'theme': user_settings.theme})
            
            # Check if school year changed
            if old_school_year != user_settings.school_year:
                messages.success(request, f'Settings updated successfully! School year changed to {user_settings.school_year}. All data will now filter by this academic year.')
            else:
                messages.success(request, 'Settings updated successfully!')
        except Exception as e:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, 'Please run database migration first: python manage.py migrate')
        return redirect('settings')

    return render(request, 'settings.html', {'user_settings': user_settings})


@login_required
def upload_students_csv(request, class_id):
    class_obj = get_object_or_404(Class, id=class_id)

    if request.method == "POST" and request.FILES.get("csv_file"):
        uploaded_file = request.FILES["csv_file"]
        ext = uploaded_file.name.split(".")[-1].lower()

        try:
            if ext == "csv":
                data = uploaded_file.read().decode("utf-8-sig")
                csv_reader = csv.reader(io.StringIO(data))
                headers = next(csv_reader)
                rows = list(csv_reader)
            elif ext in ["xlsx", "xls"]:
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                ws = wb.active
                headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = list(ws.iter_rows(min_row=2, values_only=True))
            else:
                messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
                return render(request, "upload_csv.html", {"class_obj": class_obj})
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return render(request, "upload_csv.html", {"class_obj": class_obj})

        # Normalize headers
        headers_normalized = [h.strip().lower().replace("_", " ") for h in headers]

        # Allowed variations (in case CSV uses 'middlename' instead of 'middle name')
        header_aliases = {
            "middlename": "middle name",
            "middle": "middle name",
            "middle initial": "middle name",
        }
        headers_normalized = [header_aliases.get(h, h) for h in headers_normalized]

        required_headers = [
            "code", "last name", "first name", "middle name",
            "sex", "course", "year", "units", "section"
        ]

        # Identify missing/unexpected headers for better debugging
        missing = [h for h in required_headers if h not in headers_normalized]
        unexpected = [h for h in headers_normalized if h not in required_headers]

        if missing or unexpected:
            msg = "❌ CSV header mismatch.<br>"
            if missing:
                msg += f"<b>Missing:</b> {', '.join(missing)}<br>"
            if unexpected:
                msg += f"<b>Unexpected:</b> {', '.join(unexpected)}"
            messages.error(request, msg)
            return render(request, "upload_csv.html", {"class_obj": class_obj})

        success_count = 0
        for row in rows:
            if not any(row):
                continue

            try:
                data = dict(zip(headers_normalized, row))

                # Create or update student record
                Student.objects.update_or_create(
                    student_id=str(data.get("code", "")).strip(),
                    defaults={
                        "last_name": str(data.get("last name", "")).strip(),
                        "first_name": str(data.get("first name", "")).strip(),
                        "middle_initial": str(data.get("middle name", "") or "").strip(),
                        "program": str(data.get("course", "")).strip(),
                        "year_level": str(data.get("year", "")).strip(),
                        "section": str(data.get("section", "")).strip(),
                        "class_obj": class_obj,
                    },
                )
                success_count += 1
            except Exception as e:
                messages.warning(request, f"⚠️ Skipped a row due to error: {e}")

        messages.success(request, f"✅ Successfully imported {success_count} students.")
        return redirect("class_detail", class_id=class_obj.id)

    return render(request, "upload_csv.html", {"class_obj": class_obj})


@login_required
def class_panel(request):
    current_school_year = get_current_school_year(request.user)
    classes = Class.objects.filter(instructor=request.user, school_year=current_school_year)
    if request.method == "POST":
        form = ClassForm(request.POST, user=request.user)
        if form.is_valid():
            new_class = form.save(commit=False)
            new_class.instructor = request.user
            new_class.save()
            messages.success(request, "Class added successfully!")
            return redirect("class_panel")
    else:
        form = ClassForm(user=request.user)
    context = {
        "class_form": form, 
        "classes": classes,
        "user_theme": get_user_theme(request.user)
    }
    return render(request, "class.html", context)


# -----------------------------
# ACTIVITY LOG / ABOUT / REPORTS (unchanged)
# -----------------------------
@login_required
def activity_log(request):
    """View recent activity log"""
    activities = ActivityLog.objects.filter(user=request.user)[:50]
    return render(request, 'activity_log.html', {'activities': activities})


def about_view(request):
    """About page"""
    return render(request, 'about.html')


@login_required
def generate_attendance_report(request, class_id):
    """Generate attendance report for a class"""
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    students = Student.objects.filter(class_obj=class_obj)

    # Calculate attendance statistics
    attendance_stats = []
    for student in students:
        total_days = Attendance.objects.filter(student=student, class_obj=class_obj).count()
        present_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Present').count()
        absent_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Absent').count()
        late_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Late').count()
        excused_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Excused').count()

        attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0

        attendance_stats.append({
            'student': student,
            'total_days': total_days,
            'present_days': present_days,
            'absent_days': absent_days,
            'late_days': late_days,
            'excused_days': excused_days,
            'attendance_percentage': round(attendance_percentage, 2)
        })

    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action=f"Generated attendance report for {class_obj.program}",
        description=f"PDF report generated for {len(attendance_stats)} students",
        class_obj=class_obj
    )

    return render(request, 'reports/attendance_report.html', {
        'class_obj': class_obj,
        'attendance_stats': attendance_stats,
        'generated_date': timezone.now(),
    })


@login_required
def generate_grade_report(request, class_id):
    """Generate grade report for a class"""
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    students = Student.objects.filter(class_obj=class_obj)
    summaries = GradeSummary.objects.filter(class_obj=class_obj)

    # Get grade settings
    try:
        grade_settings = class_obj.grade_settings
    except GradeCalculationSettings.DoesNotExist:
        grade_settings = GradeCalculationSettings.objects.create(class_obj=class_obj)

    # Calculate statistics
    total_students = students.count()
    passed_count = summaries.filter(remarks='Passed').count()
    failed_count = summaries.filter(remarks='Failed').count()
    average_grade = summaries.aggregate(avg_grade=Avg('final_grade'))['avg_grade'] or 0

    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action=f"Generated grade report for {class_obj.program}",
        description=f"PDF report generated for {total_students} students",
        class_obj=class_obj
    )

    return render(request, 'reports/grade_report.html', {
        'class_obj': class_obj,
        'summaries': summaries,
        'grade_settings': grade_settings,
        'total_students': total_students,
        'passed_count': passed_count,
        'failed_count': failed_count,
        'average_grade': round(average_grade, 2),
        'generated_date': timezone.now(),
    })


@login_required
def generate_class_summary(request, class_id):
    """Generate comprehensive class summary report"""
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    students = Student.objects.filter(class_obj=class_obj)
    summaries = GradeSummary.objects.filter(class_obj=class_obj)

    # Attendance statistics
    attendance_stats = []
    for student in students:
        total_days = Attendance.objects.filter(student=student, class_obj=class_obj).count()
        present_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Present').count()
        attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0

        attendance_stats.append({
            'student': student,
            'attendance_percentage': round(attendance_percentage, 2)
        })

    # Grade statistics
    total_students = students.count()
    passed_count = summaries.filter(remarks='Passed').count()
    failed_count = summaries.filter(remarks='Failed').count()
    average_grade = summaries.aggregate(avg_grade=Avg('final_grade'))['avg_grade'] or 0

    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action=f"Generated class summary for {class_obj.program}",
        description=f"Comprehensive report generated",
        class_obj=class_obj
    )

    return render(request, 'reports/class_summary.html', {
        'class_obj': class_obj,
        'students': students,
        'summaries': summaries,
        'attendance_stats': attendance_stats,
        'total_students': total_students,
        'passed_count': passed_count,
        'failed_count': failed_count,
        'average_grade': round(average_grade, 2),
        'generated_date': timezone.now(),
    })


@login_required
def reports_panel(request):
    """Main reports panel"""
    current_school_year = get_current_school_year(request.user)
    classes = Class.objects.filter(instructor=request.user, school_year=current_school_year)
    return render(request, 'reports_panel.html', {
        'classes': classes, 
        'current_school_year': current_school_year,
        'user_theme': get_user_theme(request.user)
    })


# PDF Report Generation (HTML format)
@login_required
def generate_attendance_pdf(request, class_id):
    """Generate attendance report as HTML (printable)"""
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    students = Student.objects.filter(class_obj=class_obj)
    
    attendance_stats = []
    for student in students:
        total_days = Attendance.objects.filter(student=student, class_obj=class_obj).count()
        present_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Present').count()
        absent_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Absent').count()
        late_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Late').count()
        excused_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Excused').count()
        attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0
        
        attendance_stats.append({
            'student': student,
            'total_days': total_days,
            'present_days': present_days,
            'absent_days': absent_days,
            'late_days': late_days,
            'excused_days': excused_days,
            'attendance_percentage': round(attendance_percentage, 2)
        })
    
    return render(request, 'reports/attendance_pdf.html', {
        'class_obj': class_obj,
        'attendance_stats': attendance_stats,
        'generated_date': timezone.now(),
    })


@login_required
def generate_grades_pdf(request, class_id):
    """Generate grades report as HTML (printable)"""
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    summaries = GradeSummary.objects.filter(class_obj=class_obj).order_by('student__last_name')
    
    return render(request, 'reports/grades_pdf.html', {
        'class_obj': class_obj,
        'summaries': summaries,
        'generated_date': timezone.now(),
    })


@login_required
def generate_summary_pdf(request, class_id):
    """Generate class summary report as HTML (printable)"""
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    students = Student.objects.filter(class_obj=class_obj)
    summaries = GradeSummary.objects.filter(class_obj=class_obj)
    
    # Attendance stats
    attendance_stats = []
    for student in students:
        total_days = Attendance.objects.filter(student=student, class_obj=class_obj).count()
        present_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Present').count()
        attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0
        attendance_stats.append({
            'student': student,
            'attendance_percentage': round(attendance_percentage, 2)
        })
    
    return render(request, 'reports/summary_pdf.html', {
        'class_obj': class_obj,
        'students': students,
        'summaries': summaries,
        'attendance_stats': attendance_stats,
        'generated_date': timezone.now(),
    })


# Excel Report Generation
@login_required
def generate_attendance_excel(request, class_id):
    """Generate attendance report as Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    students = Student.objects.filter(class_obj=class_obj)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Headers
    headers = ['Student ID', 'Student Name', 'Total Days', 'Present', 'Absent', 'Late', 'Excused', 'Attendance %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, student in enumerate(students, 2):
        total_days = Attendance.objects.filter(student=student, class_obj=class_obj).count()
        present_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Present').count()
        absent_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Absent').count()
        late_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Late').count()
        excused_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Excused').count()
        attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0
        
        ws.cell(row=row, column=1, value=student.student_id)
        ws.cell(row=row, column=2, value=student.display_name)
        ws.cell(row=row, column=3, value=total_days)
        ws.cell(row=row, column=4, value=present_days)
        ws.cell(row=row, column=5, value=absent_days)
        ws.cell(row=row, column=6, value=late_days)
        ws.cell(row=row, column=7, value=excused_days)
        ws.cell(row=row, column=8, value=f"{attendance_percentage:.2f}%")
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{class_obj.program}.xlsx"'
    wb.save(response)
    return response


@login_required
def generate_grades_excel(request, class_id):
    """Generate grades report as Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    summaries = GradeSummary.objects.filter(class_obj=class_obj).order_by('student__last_name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades Report"
    
    # Headers
    headers = ['Student ID', 'Student Name', 'Final Grade', 'Equivalent Grade', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, summary in enumerate(summaries, 2):
        ws.cell(row=row, column=1, value=summary.student.student_id)
        ws.cell(row=row, column=2, value=summary.student.display_name)
        ws.cell(row=row, column=3, value=f"{summary.final_grade:.2f}%")
        ws.cell(row=row, column=4, value=summary.equivalent_grade or '-')
        ws.cell(row=row, column=5, value=summary.remarks)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="grades_report_{class_obj.program}.xlsx"'
    wb.save(response)
    return response


@login_required
def generate_summary_excel(request, class_id):
    """Generate class summary report as Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    class_obj = get_object_or_404(Class, id=class_id, instructor=request.user)
    students = Student.objects.filter(class_obj=class_obj)
    summaries = {s.student.id: s for s in GradeSummary.objects.filter(class_obj=class_obj)}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Class Summary"
    
    # Headers
    headers = ['Student ID', 'Student Name', 'Final Grade', 'Equivalent Grade', 'Remarks', 'Attendance %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, student in enumerate(students, 2):
        summary = summaries.get(student.id)
        
        # Calculate attendance
        total_days = Attendance.objects.filter(student=student, class_obj=class_obj).count()
        present_days = Attendance.objects.filter(student=student, class_obj=class_obj, status='Present').count()
        attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0
        
        ws.cell(row=row, column=1, value=student.student_id)
        ws.cell(row=row, column=2, value=student.display_name)
        ws.cell(row=row, column=3, value=f"{summary.final_grade:.2f}%" if summary else '-')
        ws.cell(row=row, column=4, value=summary.equivalent_grade if summary and summary.equivalent_grade else '-')
        ws.cell(row=row, column=5, value=summary.remarks if summary else '-')
        ws.cell(row=row, column=6, value=f"{attendance_percentage:.2f}%")
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="summary_report_{class_obj.program}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_all_data(request):
    """Export all user data to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    wb = openpyxl.Workbook()
    
    # Classes sheet
    ws_classes = wb.active
    ws_classes.title = "Classes"
    classes_headers = ['Program', 'Subject', 'Year Level', 'Section', 'Semester', 'School Year']
    for col, header in enumerate(classes_headers, 1):
        cell = ws_classes.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    classes = Class.objects.filter(instructor=request.user)
    for row, cls in enumerate(classes, 2):
        ws_classes.cell(row=row, column=1, value=cls.program)
        ws_classes.cell(row=row, column=2, value=cls.subject)
        ws_classes.cell(row=row, column=3, value=cls.year_level)
        ws_classes.cell(row=row, column=4, value=cls.section)
        ws_classes.cell(row=row, column=5, value=cls.semester)
        ws_classes.cell(row=row, column=6, value=cls.school_year)
    
    # Students sheet
    ws_students = wb.create_sheet("Students")
    students_headers = ['Student ID', 'Last Name', 'First Name', 'Middle Initial', 'Program', 'Year Level', 'Section', 'Class']
    for col, header in enumerate(students_headers, 1):
        cell = ws_students.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    students = Student.objects.filter(class_obj__instructor=request.user)
    for row, student in enumerate(students, 2):
        ws_students.cell(row=row, column=1, value=student.student_id)
        ws_students.cell(row=row, column=2, value=student.last_name)
        ws_students.cell(row=row, column=3, value=student.first_name)
        ws_students.cell(row=row, column=4, value=student.middle_initial or '')
        ws_students.cell(row=row, column=5, value=student.program)
        ws_students.cell(row=row, column=6, value=student.year_level)
        ws_students.cell(row=row, column=7, value=student.section)
        ws_students.cell(row=row, column=8, value=student.class_obj.program)
    
    # Grades sheet
    ws_grades = wb.create_sheet("Grades")
    grades_headers = ['Student ID', 'Student Name', 'Class', 'Final Grade', 'Equivalent Grade', 'Remarks']
    for col, header in enumerate(grades_headers, 1):
        cell = ws_grades.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    grades = GradeSummary.objects.filter(class_obj__instructor=request.user)
    for row, grade in enumerate(grades, 2):
        ws_grades.cell(row=row, column=1, value=grade.student.student_id)
        ws_grades.cell(row=row, column=2, value=grade.student.display_name)
        ws_grades.cell(row=row, column=3, value=grade.class_obj.program)
        ws_grades.cell(row=row, column=4, value=f"{grade.final_grade:.2f}%")
        ws_grades.cell(row=row, column=5, value=grade.equivalent_grade or '-')
        ws_grades.cell(row=row, column=6, value=grade.remarks)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ascrem_data_export.xlsx"'
    wb.save(response)
    return response
